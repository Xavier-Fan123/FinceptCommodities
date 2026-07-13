"""CSV/XLSX serialization for LPG API query rows."""

from __future__ import annotations

import csv
import io
import json
from typing import Any, Dict, Iterable, List, Tuple


def _cell(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return value


def normalize_rows(rows: Iterable[Dict[str, Any]]) -> Tuple[List[str], List[Dict[str, Any]]]:
    materialized = [dict(row) for row in rows]
    fields: List[str] = []
    seen = set()
    for row in materialized:
        for key in row:
            if key not in seen:
                seen.add(key)
                fields.append(key)
    return fields, materialized


def to_csv(rows: Iterable[Dict[str, Any]]) -> bytes:
    fields, materialized = normalize_rows(rows)
    out = io.StringIO(newline="")
    writer = csv.DictWriter(out, fieldnames=fields, extrasaction="ignore")
    if fields:
        writer.writeheader()
        for row in materialized:
            writer.writerow({key: _cell(row.get(key)) for key in fields})
    return ("\ufeff" + out.getvalue()).encode("utf-8")


def to_xlsx(rows: Iterable[Dict[str, Any]], sheet_name: str = "LPG Export") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    fields, materialized = normalize_rows(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "LPG Export")[:31]
    if fields:
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="263238")
        for row in materialized:
            ws.append([_cell(row.get(key)) for key in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for index, field in enumerate(fields, start=1):
            sample = [str(_cell(row.get(field)) or "") for row in materialized[:200]]
            ws.column_dimensions[ws.cell(1, index).column_letter].width = min(60, max(10, len(field) + 2, *(len(v) + 2 for v in sample)))
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
