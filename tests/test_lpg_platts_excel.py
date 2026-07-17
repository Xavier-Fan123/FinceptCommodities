import unittest

from openpyxl import load_workbook

from _support import WorkspaceScratchMixin
from lpg.platts_excel import (
    build_backfill_workbooks,
    build_curve_workbooks,
    build_moc_workbooks,
    build_probe_workbook,
    build_scope_workbooks,
    build_workbook,
    parse_workbook,
    write_staging,
)


class PlattsWorkbookTests(WorkspaceScratchMixin, unittest.TestCase):
    @staticmethod
    def manifest_rows(path):
        workbook = load_workbook(path, data_only=False, read_only=True)
        try:
            manifest = list(workbook["_query_manifest"].iter_rows(values_only=True))
            headers = list(manifest[0])
            return workbook.sheetnames, [
                dict(zip(headers, row)) for row in manifest[1:] if row[0]
            ], {
                row[1]: workbook[row[1]]["A2"].value for row in manifest[1:] if row[0]
            }
        finally:
            workbook.close()

    def test_probe_workbook_has_one_current_query_and_no_discovery(self):
        from unittest.mock import patch

        with patch("lpg.platts_excel.PRIVATE_DIR", self.scratch):
            path = build_probe_workbook(force=True)
        workbook = load_workbook(path, data_only=False, read_only=True)
        try:
            manifest = list(workbook["_query_manifest"].iter_rows(values_only=True))
            headers = list(manifest[0])
            rows = [dict(zip(headers, row)) for row in manifest[1:] if row[0]]
            self.assertEqual(1, len(rows))
            self.assertEqual(("cfr_na_propane", "current"),
                             (rows[0]["candidate_id"], rows[0]["kind"]))
            self.assertNotIn("datasets", workbook.sheetnames)
            self.assertNotIn("md_catalog", workbook.sheetnames)
        finally:
            workbook.close()

    def test_build_workbook_isolated_queries_and_manifest(self):
        path = build_workbook(
            self.scratch / "Platts_LPG_asia.xlsx", scope="asia", force=True,
        )
        workbook = load_workbook(path, data_only=False, read_only=True)
        try:
            self.assertEqual("veryHidden", workbook["_query_manifest"].sheet_state)
            self.assertEqual("veryHidden", workbook["_runtime_status"].sheet_state)
            manifest = list(workbook["_query_manifest"].iter_rows(values_only=True))
            headers = list(manifest[0])
            rows = [dict(zip(headers, row)) for row in manifest[1:] if row[0]]
            self.assertGreater(len(rows), 20)
            self.assertEqual(len(rows), len({row["sheet"] for row in rows}))
            cfr = [row for row in rows if row["candidate_id"] == "cfr_na_propane"]
            self.assertEqual({"current", "history", "correction"},
                             {row["kind"] for row in cfr})
            self.assertTrue(all(row["canonical_key"] == "FEI_PROPANE" for row in cfr))
            for row in rows:
                formula = workbook[row["sheet"]]["A2"].value
                self.assertIsInstance(formula, str)
                self.assertTrue(formula.startswith("=PlattsGetData("))
        finally:
            workbook.close()

    def test_daily_scope_workbook_contains_only_current_market_queries(self):
        from unittest.mock import patch

        with patch("lpg.platts_excel.PRIVATE_DIR", self.scratch):
            path = build_scope_workbooks(scope="asia", force=True)[0]

        sheetnames, rows, _ = self.manifest_rows(path)
        self.assertTrue(rows)
        self.assertEqual({"current"}, {row["kind"] for row in rows})
        self.assertNotIn("datasets", sheetnames)
        self.assertFalse(any(row["dataset"] in {"FC", "eWMD"} for row in rows))

    def test_history_backfill_batches_use_assessdate_and_bate(self):
        from unittest.mock import patch

        with patch("lpg.platts_excel.PRIVATE_DIR", self.scratch):
            paths = build_backfill_workbooks(
                start_year=2025,
                end_year=2025,
                scope="asia",
                symbols=("PMAAV00", "PMAAF00"),
                batch_size=1,
                force=True,
            )

        self.assertEqual(2, len(paths))
        for path in paths:
            _, rows, formulas = self.manifest_rows(path)
            self.assertEqual(1, len(rows))
            self.assertEqual("history", rows[0]["kind"])
            formula = formulas[rows[0]["sheet"]]
            self.assertIn("Symbol in ('", formula)
            self.assertIn("AssessDate>=2025-01-01", formula)
            self.assertNotIn("AssessDate<=", formula)
            self.assertIn("Bate in ('c','u')", formula)
            self.assertNotIn("ModDate", formula)

    def test_curve_workbook_uses_curve_selection_not_market_candidate_filter(self):
        from unittest.mock import patch

        with patch("lpg.platts_excel.PRIVATE_DIR", self.scratch):
            paths = build_curve_workbooks(
                scope="asia", curve_ids=("CN3HO",), force=True,
            )

        self.assertEqual(1, len(paths))
        _, rows, formulas = self.manifest_rows(paths[0])
        self.assertEqual(1, len(rows))
        self.assertEqual(
            ("curve_cfr_na_propane", "FC", "curve", "CN3HO"),
            (rows[0]["candidate_id"], rows[0]["dataset"], rows[0]["kind"],
             rows[0]["curve_code"]),
        )
        self.assertIn("CurveCode='CN3HO'", formulas[rows[0]["sheet"]])

    def test_moc_workbook_contains_only_ewindow_trade_query(self):
        from unittest.mock import patch

        with patch("lpg.platts_excel.PRIVATE_DIR", self.scratch):
            path = build_moc_workbooks(scope="asia", force=True)[0]

        _, rows, _ = self.manifest_rows(path)
        self.assertEqual(1, len(rows))
        self.assertEqual(("eWMD", "TradeData", "ewindow"),
                         (rows[0]["dataset"], rows[0]["data_series"], rows[0]["kind"]))

    def test_overnight_moc_query_is_restricted_to_americas(self):
        from unittest.mock import patch

        with patch("lpg.platts_excel.PRIVATE_DIR", self.scratch):
            path = build_moc_workbooks(scope="overnight", force=True)[0]

        _, rows, formulas = self.manifest_rows(path)
        self.assertEqual(1, len(rows))
        formula = formulas[rows[0]["sheet"]]
        self.assertIn("Window_Region='americas'", formula)

    def test_curve_contract_code_is_also_a_contract_label(self):
        from unittest.mock import patch

        with patch("lpg.platts_excel.PRIVATE_DIR", self.scratch):
            path = build_curve_workbooks(
                scope="asia", curve_ids=("CN3HO",), force=True,
            )[0]
        workbook = load_workbook(path, data_only=False)
        manifest = workbook["_query_manifest"]
        headers = [cell.value for cell in manifest[1]]
        spec = dict(zip(headers, next(manifest.iter_rows(min_row=2, values_only=True))))
        query = workbook[spec["sheet"]]
        query.delete_rows(1, query.max_row)
        query.append(["CurveCode", "ContractCode", "Date", "Value", "Currency", "UOM"])
        query.append(["CN3HO", "AAXXX01", "2026-07-13", 512.5, "USD", "MT"])
        runtime = workbook["_runtime_status"]
        for row in runtime.iter_rows(min_row=2):
            if row[0].value == "refresh_state":
                row[1].value = "success"
        workbook.save(path)
        workbook.close()

        payload = parse_workbook(path)

        self.assertEqual("curve", payload["purpose"])
        self.assertEqual(1, len(payload["records"]))
        self.assertEqual(
            ("AAXXX01", "AAXXX01"),
            (payload["records"][0]["symbol"], payload["records"][0]["contract_label"]),
        )

    def test_generic_parse_stages_backfill_away_from_daily_latest(self):
        from unittest.mock import patch

        with patch("lpg.platts_excel.PRIVATE_DIR", self.scratch):
            path = build_backfill_workbooks(
                start_year=2025, end_year=2025, scope="asia",
                symbols=("PMAAV00",), force=True,
            )[0]
        workbook = load_workbook(path, data_only=False)
        manifest = workbook["_query_manifest"]
        headers = [cell.value for cell in manifest[1]]
        spec = dict(zip(headers, next(manifest.iter_rows(min_row=2, values_only=True))))
        query = workbook[spec["sheet"]]
        query.delete_rows(1, query.max_row)
        query.append(["Symbol", "Value", "BATE", "Assess Date"])
        query.append(["PMAAV00", 520.0, "c", "2025-07-13"])
        query.append(["PMAAV00", 540.0, "c", "2026-07-13"])
        runtime = workbook["_runtime_status"]
        for row in runtime.iter_rows(min_row=2):
            if row[0].value == "refresh_state":
                row[1].value = "success"
        workbook.save(path)
        workbook.close()

        payload = parse_workbook(path)
        stage_dir = self.scratch / "staging"
        staged = write_staging(payload, staging_dir=stage_dir)

        self.assertEqual(("backfill", 2025), (payload["purpose"], payload["year"]))
        self.assertEqual(
            ["2025-07-13"],
            [record["assess_date"] for record in payload["records"]],
        )
        self.assertEqual(stage_dir / "backfill", staged.parent)
        self.assertTrue(staged.name.startswith("2025_asia_"))
        self.assertFalse((stage_dir / "latest.json").exists())

    def test_parse_workbook_fixture_maps_market_data_and_discovery(self):
        path = build_workbook(
            self.scratch / "Platts_LPG_asia.xlsx", scope="asia", force=True,
        )
        workbook = load_workbook(path, data_only=False)
        manifest = workbook["_query_manifest"]
        headers = [cell.value for cell in manifest[1]]
        target = None
        for values in manifest.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            if row.get("candidate_id") == "cfr_na_propane" and row.get("kind") == "current":
                target = row
                break
        self.assertIsNotNone(target)
        query = workbook[target["sheet"]]
        query.delete_rows(1, query.max_row)
        query.append(["Symbol", "Description", "Value", "Currency", "UOM",
                      "BATE", "Assess Date", "Mod Date"])
        query.append(["PMAAV00", "CFR North Asia propane", 523.75, "USD", "MT",
                      "c", "2026-07-09", "2026-07-09T14:30:00+00:00"])
        runtime = workbook["_runtime_status"]
        for row in runtime.iter_rows(min_row=2):
            if row[0].value == "refresh_state":
                row[1].value = "success"
        discovery = workbook["md_catalog"]
        discovery["A2"] = "MD"
        discovery["B2"] = "Current-Symbol"
        workbook.save(path)
        workbook.close()

        payload = parse_workbook(path)
        records = [row for row in payload["records"]
                   if row.get("series_id") == "platts_cfr_na_propane"]
        entitlement = next(
            row for row in payload["entitlement_results"]
            if row["candidate_id"] == "cfr_na_propane"
            and row["data_series"] == "Current-Symbol"
        )

        self.assertEqual("success", payload["status"])
        self.assertEqual("success", payload["runtime_status"]["refresh_state"])
        self.assertEqual(1, len(records))
        self.assertEqual(
            ("FEI_PROPANE", "PMAAV00", 523.75, "c", "2026-07-09"),
            (records[0]["canonical_key"], records[0]["symbol"], records[0]["value"],
             records[0]["bate"], records[0]["assess_date"]),
        )
        self.assertEqual(("entitled", 1, "data_returned"),
                         (entitlement["status"], entitlement["record_count"],
                          entitlement["reason_code"]))
        catalog = next(row for row in payload["discovery"] if row["sheet"] == "md_catalog")
        self.assertEqual([["MD", "Current-Symbol"]], catalog["rows"])


if __name__ == "__main__":
    unittest.main()
