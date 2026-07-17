"""S&P Global Energy Excel Add-in workbook and staging helpers.

This module deliberately does not authenticate to Platts.  It writes only official
Excel Add-in UDFs and parses the cached values after a signed-in Excel session has
calculated and saved the workbook.  A failed calculation never replaces the
last-good JSON staging file.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import json
import math
import os
import re
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence


SCHEMA_VERSION = 1
ROOT = Path(__file__).resolve().parents[1]
PRIVATE_DIR = ROOT / "data" / "private" / "platts"
STAGING_DIR = PRIVATE_DIR / "staging"

VALID_SCOPES = {"asia", "overnight", "all"}
VALID_PURPOSES = {"daily", "backfill", "curve", "moc"}
ERROR_VALUES = {"#VALUE!", "#N/A", "#REF!", "#NAME?", "#NUM!", "#NULL!", "#DIV/0!"}
BATE_ORDER = {"c": 0, "u": 1, "e": 2, "l": 3, "h": 4}
SYMBOL_RE = re.compile(r"^[A-Z0-9]{5,10}$")
# These otherwise-valid public codes returned entitlement errors in the live
# signed-in account on 2026-07-13. Keeping them out of the daily calculation is
# essential: the Add-in retries failed UDFs indefinitely and can prevent every
# successful price in the workbook from being saved. They remain in the catalog
# and status views so the data gap is explicit.
LIVE_DAILY_EXCLUDED_SYMBOLS = {"AAHHG00", "AAHHH00", "AAHHI00"}
DISCOVERY_SHEETS = (
    "datasets", "md_catalog", "md_current_schema", "md_history_schema",
    "md_correction_schema", "md_lpg_metadata", "fc_catalog",
    "fc_curve_schema", "fc_pivot_schema", "ewmd_catalog", "ewmd_trade_schema",
)


@dataclass(frozen=True)
class Candidate:
    candidate_id: str
    label: str
    scope: str
    query_group: str
    basis: str
    search_terms: str
    symbol: str = ""
    currency: str = "USD"
    uom: str = "MT"
    enabled: bool = True

    @property
    def series_id(self) -> str:
        return f"platts_{self.candidate_id}"


@dataclass(frozen=True)
class CurveCandidate:
    candidate_id: str
    label: str
    scope: str
    query_group: str
    basis: str
    curve_code: str
    search_terms: str
    currency: str = "USD"
    uom: str = "MT"
    enabled: bool = True

    @property
    def series_id(self) -> str:
        return f"platts_{self.candidate_id}"


# Codes below are public Platts price/curve codes.  Empty codes intentionally stay
# pending until the account's Add-in search/catalog confirms the current symbol.
CANDIDATES: tuple[Candidate, ...] = (
    Candidate("saudi_cp_propane_actual", "Saudi Aramco propane CP", "asia", "cp", "monthly", "Saudi Aramco propane CP", "PTAAM10"),
    Candidate("saudi_cp_butane_actual", "Saudi Aramco butane CP", "asia", "cp", "monthly", "Saudi Aramco butane CP"),
    Candidate("saudi_cp_propane_m1", "Saudi CP propane swap M1", "asia", "cp", "Singapore close", "Propane Saudi CP Swap Mo01", "AAHHG00"),
    Candidate("saudi_cp_propane_m2", "Saudi CP propane swap M2", "asia", "cp", "Singapore close", "Propane Saudi CP Swap Mo02", "AAHHH00"),
    Candidate("saudi_cp_propane_m3", "Saudi CP propane swap M3", "asia", "cp", "Singapore close", "Propane Saudi CP Swap Mo03", "AAHHI00"),
    Candidate("cfr_na_propane", "CFR North Asia propane", "asia", "north_asia", "Singapore close", "Propane CFR North Asia refrigerated cargo", "PMAAV00"),
    Candidate("cfr_na_butane", "CFR North Asia butane", "asia", "north_asia", "Singapore close", "Butane CFR North Asia refrigerated cargo", "PMAAF00"),
    Candidate("cfr_na_propane_hm1", "CFR North Asia propane front half-month", "asia", "north_asia", "Singapore close", "CFR North Asia propane first half-month", "AAVAK00"),
    Candidate("cfr_na_propane_hm2", "CFR North Asia propane second half-month", "asia", "north_asia", "Singapore close", "CFR North Asia propane second half-month", "AAVAL00"),
    Candidate("cfr_na_propane_hm3", "CFR North Asia propane third half-month", "asia", "north_asia", "Singapore close", "CFR North Asia propane third half-month", "AAVAM00"),
    Candidate("cfr_na_butane_hm1", "CFR North Asia butane front half-month", "asia", "north_asia", "Singapore close", "CFR North Asia butane first half-month", "AAVAN00"),
    Candidate("cfr_na_butane_hm2", "CFR North Asia butane second half-month", "asia", "north_asia", "Singapore close", "CFR North Asia butane second half-month", "AAVAO00"),
    Candidate("cfr_na_butane_hm3", "CFR North Asia butane third half-month", "asia", "north_asia", "Singapore close", "CFR North Asia butane third half-month", "AAVAP00"),
    Candidate("cfr_na_1111", "CFR North Asia LPG 11:11", "asia", "north_asia", "Singapore close", "CFR North Asia refrigerated LPG 11:11", "AASGN00"),
    Candidate("fob_ag_propane_cp_diff", "FOB Arab Gulf propane vs CP", "asia", "middle_east", "Singapore close", "Propane FOB AG differential Saudi CP", "PMABF00"),
    Candidate("fob_ag_butane_cp_diff", "FOB Arab Gulf butane vs CP", "asia", "middle_east", "Singapore close", "Butane FOB AG differential Saudi CP", "PMABG00"),
    Candidate("fob_east_china_lpg", "FOB East China LPG", "asia", "china", "Singapore close", "LPG FOB East China", "AAWUZ00"),
    Candidate("fob_east_china_cp_diff", "FOB East China LPG vs CP", "asia", "china", "Singapore close", "LPG FOB East China premium Saudi CP", "AAWVA00"),
    Candidate("fob_singapore_lpg", "FOB Singapore LPG", "asia", "southeast_asia", "Singapore close", "LPG FOB Singapore", "AAWVD00"),
    Candidate("fob_singapore_cp_diff", "FOB Singapore LPG vs CP", "asia", "southeast_asia", "Singapore close", "LPG FOB Singapore premium Saudi CP", "AAWVE00"),
    Candidate("cfr_philippines_lpg", "CFR Philippines LPG", "asia", "southeast_asia", "Singapore close", "LPG CFR Philippines pressurized"),
    Candidate("cfr_vietnam_lpg", "CFR Vietnam LPG", "asia", "southeast_asia", "Singapore close", "LPG CFR Vietnam pressurized"),
    Candidate("pressurized_asia_propane", "Pressurized Asia propane", "asia", "pressurized", "Singapore close", "Propane pressurized Asia"),
    Candidate("mopj_naphtha", "Naphtha C+F Japan cargo", "asia", "naphtha", "Singapore close", "Naphtha C+F Japan cargo", "PAAAD00"),
    Candidate("vlgc_ag_japan", "VLGC Arab Gulf-Japan freight", "asia", "freight", "London close", "VLGC Arab Gulf Japan LPG freight"),
    Candidate("mb_propane_enterprise_usdmt", "Propane Enterprise Mt Belvieu USD/mt M1", "overnight", "mont_belvieu", "US close", "Propane Enterprise Mt Belvieu $/mt Mo01", "AAXDD00"),
    Candidate("mb_butane_enterprise_usdmt", "Butane Enterprise Mt Belvieu USD/mt M1", "overnight", "mont_belvieu", "US close", "Butane Enterprise Mt Belvieu $/mt Mo01", "AAXDC00"),
    Candidate("mb_propane_enterprise_cpg", "Propane Enterprise Mt Belvieu M1", "overnight", "mont_belvieu", "US close", "Propane Enterprise Mt Belvieu Mo01", "PMAAY00", "USC", "GAL"),
    Candidate("mb_propane_energy_transfer_cpg", "Propane Energy Transfer Mt Belvieu M1", "overnight", "mont_belvieu", "US close", "Propane Energy Transfer Mt Belvieu Mo01", "PMABQ00", "USC", "GAL"),
    Candidate("mb_butane_enterprise_cpg", "Butane Enterprise Mt Belvieu M1", "overnight", "mont_belvieu", "US close", "Butane Enterprise Mt Belvieu Mo01", "PMAAI00", "USC", "GAL"),
    Candidate("mb_butane_energy_transfer_cpg", "Butane Energy Transfer Mt Belvieu", "overnight", "mont_belvieu", "US close", "Butane Energy Transfer Mt Belvieu", "PMABR00", "USC", "GAL"),
    Candidate("fob_usgc_lpg_2222", "LPG 22:22 FOB USGC", "overnight", "usgc", "US close", "LPG 22:22 FOB USGC $/mt"),
    Candidate("vlgc_houston_chiba", "VLGC Houston-Chiba freight", "overnight", "freight", "London close", "VLGC Houston Chiba LPG freight"),
)

CURVE_CANDIDATES: tuple[CurveCandidate, ...] = (
    CurveCandidate("curve_saudi_cp_propane", "Saudi CP propane derivatives", "asia", "cp", "Singapore close", "CN0PT", "Saudi CP propane derivatives"),
    CurveCandidate("curve_cfr_na_propane", "CFR North Asia propane derivatives", "asia", "north_asia", "Singapore close", "CN3HO", "CFR North Asia propane derivatives"),
    CurveCandidate("curve_fei_cp", "CFR North Asia vs Saudi CP propane", "asia", "spreads", "Singapore close", "CN3HP", "CFR North Asia FOB Saudi CP propane derivatives"),
    CurveCandidate("curve_fei_mopj", "CFR North Asia propane vs MOPJ", "asia", "spreads", "Singapore close", "CN3HQ", "CFR North Asia propane versus naphtha MOPJ"),
)


# Keep the workbook-oriented candidate ids stable while mapping them onto the
# canonical keys used by the analytics layer.  Unmapped discovery targets are
# still retained with an upper-case key; they simply do not participate in a
# derived spread until their semantic mapping has been reviewed.
CANONICAL_KEYS: dict[str, str] = {
    "saudi_cp_propane_actual": "CP_PROPANE",
    "saudi_cp_butane_actual": "CP_BUTANE",
    "saudi_cp_propane_m1": "CP_PROPANE_M1",
    "saudi_cp_propane_m2": "CP_PROPANE_M2",
    "saudi_cp_propane_m3": "CP_PROPANE_M3",
    "cfr_na_propane": "FEI_PROPANE",
    "cfr_na_butane": "CFR_NA_BUTANE",
    "cfr_na_propane_hm1": "FEI_PROPANE_HM1",
    "cfr_na_propane_hm2": "FEI_PROPANE_HM2",
    "cfr_na_propane_hm3": "FEI_PROPANE_HM3",
    "cfr_na_butane_hm1": "CFR_NA_BUTANE_HM1",
    "cfr_na_butane_hm2": "CFR_NA_BUTANE_HM2",
    "cfr_na_butane_hm3": "CFR_NA_BUTANE_HM3",
    "cfr_na_1111": "CFR_NA_LPG_1111",
    "fob_ag_propane_cp_diff": "FOB_AG_PROPANE_CP_DIFF",
    "fob_ag_butane_cp_diff": "FOB_AG_BUTANE_CP_DIFF",
    "fob_east_china_lpg": "FOB_EAST_CHINA_LPG",
    "fob_east_china_cp_diff": "FOB_EAST_CHINA_CP_DIFF",
    "fob_singapore_lpg": "FOB_SINGAPORE_LPG",
    "fob_singapore_cp_diff": "FOB_SINGAPORE_CP_DIFF",
    "cfr_philippines_lpg": "CFR_PHILIPPINES_LPG",
    "cfr_vietnam_lpg": "CFR_VIETNAM_LPG",
    "pressurized_asia_propane": "PRESSURIZED_ASIA_PROPANE",
    "mopj_naphtha": "MOPJ_NAPHTHA",
    "vlgc_ag_japan": "VLGC_AG_JAPAN",
    "mb_propane_enterprise_usdmt": "MB_PROPANE_USD_MT",
    "mb_butane_enterprise_usdmt": "MB_BUTANE_USD_MT",
    "mb_propane_enterprise_cpg": "MB_PROPANE_ENTERPRISE_USC_GAL",
    "mb_propane_energy_transfer_cpg": "MB_PROPANE_ET_USC_GAL",
    "mb_butane_enterprise_cpg": "MB_BUTANE_ENTERPRISE_USC_GAL",
    "mb_butane_energy_transfer_cpg": "MB_BUTANE_ET_USC_GAL",
    "fob_usgc_lpg_2222": "FOB_USGC_LPG_2222",
    "vlgc_houston_chiba": "VLGC_USGC_NA",
    "curve_saudi_cp_propane": "CURVE_CP_PROPANE",
    "curve_cfr_na_propane": "CURVE_FEI_PROPANE",
    "curve_fei_cp": "CURVE_FEI_CP_PROPANE",
    "curve_fei_mopj": "CURVE_FEI_MOPJ",
    "ewindow_asia_asia_lpg": "EWINDOW_LPG_TRADES",
    "ewindow_overnight_americas_ngl": "EWINDOW_AMERICAS_LPG_TRADES",
}


def canonical_key(candidate_id: str) -> str:
    return CANONICAL_KEYS.get(candidate_id, candidate_id.upper())


@dataclass(frozen=True)
class QuerySpec:
    query_id: str
    sheet: str
    candidate_id: str
    canonical_key: str
    series_id: str
    label: str
    dataset: str
    data_series: str
    kind: str
    scope: str
    query_group: str
    symbol: str = ""
    curve_code: str = ""
    basis: str = ""
    currency: str = ""
    uom: str = ""
    year: int | None = None


def candidates_for_scope(scope: str) -> tuple[Candidate, ...]:
    _validate_scope(scope)
    return tuple(c for c in CANDIDATES if scope == "all" or c.scope == scope)


def curves_for_scope(scope: str) -> tuple[CurveCandidate, ...]:
    _validate_scope(scope)
    return tuple(c for c in CURVE_CANDIDATES if scope == "all" or c.scope == scope)


def workbook_path(
    scope: str,
    *,
    purpose: str = "daily",
    year: int | None = None,
    batch_id: str | None = None,
) -> Path:
    _validate_scope(scope)
    if scope == "all":
        raise ValueError("'all' uses the asia and overnight workbooks; it has no single workbook")
    if purpose not in VALID_PURPOSES:
        raise ValueError(f"purpose must be one of {sorted(VALID_PURPOSES)}")
    suffix = f"_{_safe_slug(batch_id)}" if batch_id else ""
    if purpose == "daily":
        return PRIVATE_DIR / f"Platts_LPG_{scope}.xlsx"
    if purpose == "backfill" and year is not None:
        return PRIVATE_DIR / "backfill" / f"Platts_LPG_{scope}_{year}{suffix}.xlsx"
    if purpose == "curve" and year is None:
        return PRIVATE_DIR / "curves" / f"Platts_LPG_FC_{scope}{suffix}.xlsx"
    if purpose == "moc" and year is None:
        return PRIVATE_DIR / "moc" / f"Platts_LPG_MOC_{scope}{suffix}.xlsx"
    if purpose == "backfill":
        raise ValueError("backfill workbooks require a year")
    raise ValueError(f"{purpose} workbooks do not accept a year")


def build_workbook(
    path: Path | str,
    *,
    scope: str,
    purpose: str = "daily",
    year: int | None = None,
    force: bool = False,
    include_discovery: bool = True,
    candidate_ids: Sequence[str] | None = None,
    curve_candidate_ids: Sequence[str] | None = None,
    query_kinds: Sequence[str] | None = None,
    include_curves: bool = False,
    include_ewindow: bool = False,
) -> Path:
    """Build an Add-in workbook without starting Excel or storing credentials."""
    _validate_scope(scope)
    if scope == "all":
        raise ValueError("build one workbook per concrete scope")
    if purpose not in VALID_PURPOSES:
        raise ValueError(f"purpose must be one of {sorted(VALID_PURPOSES)}")
    if purpose == "backfill" and year is None:
        raise ValueError("backfill workbook requires year")
    if year is not None and not 1990 <= year <= 2200:
        raise ValueError("year is outside the supported range")
    selected_ids = set(candidate_ids) if candidate_ids is not None else None
    selected_curve_ids = set(curve_candidate_ids) if curve_candidate_ids is not None else None
    selected_kinds = set(
        query_kinds if query_kinds is not None else ("current", "history", "correction")
    )
    if not selected_kinds <= {"current", "history", "correction"}:
        raise ValueError("query_kinds contains an unsupported market-data query")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    target = Path(path).resolve()
    if target.exists() and not force:
        return target
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    # The Platts async UDF is volatile under Excel automatic calculation and
    # can dispatch the same HTTP request repeatedly. The desktop runner opens
    # the workbook in manual mode and calculates each manifest cell once.
    wb.calculation.calcMode = "manual"
    wb.calculation.fullCalcOnLoad = False
    wb.calculation.forceFullCalc = False
    readme = wb.create_sheet("README")
    notes = [
        "FinceptCommodities LPG - S&P Global Energy Excel Add-in workbook",
        "",
        "This workbook contains only official Platts Excel Add-in UDFs.",
        "Credentials are never stored. Sign in through the S&P Global Energy ribbon.",
        "Each market-data candidate has its own sheet so an unentitled symbol cannot break other queries.",
        "Do not combine symbols into a single Symbol in (...) expression.",
        f"Scope: {scope}",
        f"Purpose: {purpose}",
        f"Year: {year if year is not None else 'rolling'}",
        "",
        "Generated formulas use PlattsDataSet, PlattsDataSeries, PlattsGetMetaData and PlattsGetData.",
        "The refresh script preserves the saved workbook when formulas fail or the session has expired.",
    ]
    for row, line in enumerate(notes, 1):
        readme.cell(row, 1, line)
    readme.column_dimensions["A"].width = 110

    catalog = wb.create_sheet("candidate_catalog")
    catalog_headers = [
        "candidate_id", "canonical_key", "series_id", "label", "scope", "query_group", "basis",
        "search_terms", "symbol", "currency", "uom", "enabled", "initial_status",
    ]
    catalog.append(catalog_headers)
    for candidate in candidates_for_scope(scope):
        if selected_ids is not None and candidate.candidate_id not in selected_ids:
            continue
        catalog.append([
            candidate.candidate_id, canonical_key(candidate.candidate_id), candidate.series_id,
            candidate.label, candidate.scope,
            candidate.query_group, candidate.basis, candidate.search_terms, candidate.symbol,
            candidate.currency, candidate.uom, candidate.enabled,
            "pending_review" if not candidate.symbol else "pending_test",
        ])
    if include_curves:
        for curve in curves_for_scope(scope):
            if selected_curve_ids is not None and curve.candidate_id not in selected_curve_ids:
                continue
            catalog.append([
                curve.candidate_id, canonical_key(curve.candidate_id), curve.series_id,
                curve.label, curve.scope,
                curve.query_group, curve.basis, curve.search_terms, curve.curve_code,
                curve.currency, curve.uom, curve.enabled, "pending_test",
            ])

    manifest = wb.create_sheet("_query_manifest")
    manifest_headers = [f.name for f in QuerySpec.__dataclass_fields__.values()]
    manifest.append(manifest_headers)
    manifest.sheet_state = "veryHidden"

    runtime = wb.create_sheet("_runtime_status")
    runtime.append(["key", "value"])
    runtime.append(["refresh_state", "not_run"])
    runtime.append(["refreshed_at", ""])
    runtime.append(["resolved_queries", 0])
    runtime.append(["resolved_discovery", 0])
    runtime.append(["purpose", purpose])
    runtime.append(["year", year or ""])
    runtime.append(["message", "Workbook has not been calculated by the official Add-in yet."])
    runtime.sheet_state = "veryHidden"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in (catalog, manifest):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    # Schema/discovery formulas stay isolated from production query sheets.
    discovery = (
        ("datasets", '=PlattsDataSet()'),
        ("md_catalog", '=PlattsDataSet("MD")'),
        ("md_current_schema", '=PlattsDataSeries("MD","Current-Symbol")'),
        ("md_history_schema", '=PlattsDataSeries("MD","History-Symbol")'),
        ("md_correction_schema", '=PlattsDataSeries("MD","Correction-Symbol")'),
        ("md_lpg_metadata", '=PlattsGetMetaData("MD","Symbols",,"Description=\'*LPG*\' or Description=\'*Propane*\' or Description=\'*Butane*\'")'),
        ("fc_catalog", '=PlattsDataSet("FC")'),
        ("fc_curve_schema", '=PlattsDataSeries("FC","CurveData")'),
        ("fc_pivot_schema", '=PlattsDataSeries("FC","CurveData-Pivoted")'),
        ("ewmd_catalog", '=PlattsDataSet("eWMD")'),
        ("ewmd_trade_schema", '=PlattsDataSeries("eWMD","TradeData")'),
    )
    if include_discovery:
        for name, formula in discovery:
            ws = wb.create_sheet(name[:31])
            ws["A1"] = "Discovery/schema query - an error here does not affect price queries"
            ws["A2"] = formula
            ws.column_dimensions["A"].width = 100

    specs: list[tuple[QuerySpec, str]] = []
    qnum = 1

    def add_spec(spec: QuerySpec, formula: str) -> None:
        nonlocal qnum
        specs.append((spec, formula))
        qnum += 1

    for candidate in candidates_for_scope(scope):
        if selected_ids is not None and candidate.candidate_id not in selected_ids:
            continue
        if purpose in {"curve", "moc"}:
            continue
        if not candidate.enabled or not candidate.symbol:
            continue
        if purpose == "backfill":
            formula = _history_formula(candidate.symbol, year=year)
            kind = "history"
            series = "History-Symbol"
            sheet = f"q{qnum:03d}_history"
            add_spec(QuerySpec(
                f"{candidate.candidate_id}:history:{year}", sheet, candidate.candidate_id,
                canonical_key(candidate.candidate_id), candidate.series_id, candidate.label,
                "MD", series, kind, candidate.scope,
                candidate.query_group, candidate.symbol, basis=candidate.basis,
                currency=candidate.currency, uom=candidate.uom, year=year,
            ), formula)
            continue

        query_defs = (
            ("current", "Current-Symbol", _current_formula(candidate.symbol)),
            ("history", "History-Symbol", _history_formula(candidate.symbol)),
            ("correction", "Correction-Symbol", _correction_formula(candidate.symbol)),
        )
        for kind, series, formula in query_defs:
            if kind not in selected_kinds:
                continue
            sheet = f"q{qnum:03d}_{kind[:8]}"
            add_spec(QuerySpec(
                f"{candidate.candidate_id}:{kind}", sheet, candidate.candidate_id,
                canonical_key(candidate.candidate_id), candidate.series_id, candidate.label,
                "MD", series, kind, candidate.scope,
                candidate.query_group, candidate.symbol, basis=candidate.basis,
                currency=candidate.currency, uom=candidate.uom,
            ), formula)

    if include_curves:
        for curve in curves_for_scope(scope):
            if selected_curve_ids is not None and curve.candidate_id not in selected_curve_ids:
                continue
            if not curve.enabled or not curve.curve_code:
                continue
            sheet = f"q{qnum:03d}_curve"
            formula = _curve_formula(curve.curve_code, year=year if purpose == "backfill" else None)
            add_spec(QuerySpec(
                f"{curve.candidate_id}:curve:{year or 'rolling'}", sheet, curve.candidate_id,
                canonical_key(curve.candidate_id), curve.series_id, curve.label,
                "FC", "CurveData", "curve", curve.scope,
                curve.query_group, curve_code=curve.curve_code, basis=curve.basis,
                currency=curve.currency, uom=curve.uom, year=year,
            ), formula)

    ewindow_filters = _ewindow_filters(scope, purpose=purpose, year=year) if include_ewindow else ()
    for group, filter_text in ewindow_filters:
        sheet = f"q{qnum:03d}_ewindow"
        candidate_id = f"ewindow_{scope}_{group}"
        add_spec(QuerySpec(
            f"{candidate_id}:{year or 'rolling'}", sheet, candidate_id,
            canonical_key(candidate_id), f"platts_{candidate_id}",
            f"eWindow LPG trades - {group}", "eWMD", "TradeData",
            "ewindow", scope, group, basis="MOC", year=year,
        ), _ewindow_formula(filter_text))

    for spec, formula in specs:
        ws = wb.create_sheet(spec.sheet)
        ws["A1"] = f"{spec.label} | {spec.dataset}/{spec.data_series} | isolated query"
        ws["A2"] = formula
        ws.column_dimensions["A"].width = 90
        manifest.append([asdict(spec)[name] for name in manifest_headers])

    catalog.column_dimensions["A"].width = 34
    catalog.column_dimensions["B"].width = 38
    catalog.column_dimensions["C"].width = 48
    catalog.column_dimensions["G"].width = 52
    catalog.column_dimensions["H"].width = 16
    wb.calculation.fullCalcOnLoad = False
    wb.calculation.forceFullCalc = False
    wb.calculation.calcMode = "manual"
    wb.save(target)
    return target


def build_scope_workbooks(*, scope: str = "all", force: bool = False) -> list[Path]:
    _validate_scope(scope)
    scopes = ("asia", "overnight") if scope == "all" else (scope,)
    paths: list[Path] = []
    for item in scopes:
        # Daily refreshes are intentionally compact: one Current-Symbol query
        # per confirmed code. Discovery, rolling history, corrections, eWindow,
        # denied symbols, and FC curves run outside the live price workbook so
        # the Add-in cannot retry one failed UDF forever and block valid prices.
        selected = tuple(
            candidate.candidate_id for candidate in candidates_for_scope(item)
            if candidate.enabled and candidate.symbol
            and candidate.symbol not in LIVE_DAILY_EXCLUDED_SYMBOLS
        )
        paths.append(build_workbook(
            workbook_path(item), scope=item, force=force, include_discovery=False,
            candidate_ids=selected, query_kinds=("current",), include_curves=False,
            include_ewindow=False,
        ))
    return paths


def build_probe_workbook(*, symbol: str = "PMAAV00", force: bool = True) -> Path:
    """Build one current-price UDF used to validate the live Add-in session."""
    if not SYMBOL_RE.fullmatch(symbol.upper()):
        raise ValueError("probe symbol is not a valid Platts symbol")
    candidate = next((item for item in CANDIDATES if item.symbol == symbol.upper()), None)
    if candidate is None:
        raise ValueError("probe symbol is not mapped to a configured LPG candidate")
    return build_workbook(
        PRIVATE_DIR / "Platts_LPG_probe.xlsx",
        scope=candidate.scope,
        force=force,
        include_discovery=False,
        candidate_ids=(candidate.candidate_id,),
        query_kinds=("current",),
        include_curves=False,
        include_ewindow=False,
    )


def _resolve_candidate_ids(
    values: Sequence[str] | None, *, scope: str,
) -> set[str] | None:
    if values is None:
        return None
    lookup: dict[str, str] = {}
    for candidate in candidates_for_scope(scope):
        lookup[candidate.candidate_id.lower()] = candidate.candidate_id
        if candidate.symbol:
            lookup[candidate.symbol.upper()] = candidate.candidate_id
    resolved: set[str] = set()
    unknown: list[str] = []
    for raw in values:
        value = str(raw or "").strip()
        candidate_id = lookup.get(value.lower()) or lookup.get(value.upper())
        if candidate_id:
            resolved.add(candidate_id)
        else:
            unknown.append(value or "<empty>")
    if unknown:
        raise ValueError(f"unknown LPG symbol/candidate for scope {scope}: {', '.join(unknown)}")
    return resolved


def _resolve_curve_ids(
    values: Sequence[str] | None, *, scope: str,
) -> set[str] | None:
    if values is None:
        return None
    lookup: dict[str, str] = {}
    for curve in curves_for_scope(scope):
        lookup[curve.candidate_id.lower()] = curve.candidate_id
        lookup[curve.curve_code.upper()] = curve.candidate_id
    resolved: set[str] = set()
    unknown: list[str] = []
    for raw in values:
        value = str(raw or "").strip()
        curve_id = lookup.get(value.lower()) or lookup.get(value.upper())
        if curve_id:
            resolved.add(curve_id)
        else:
            unknown.append(value or "<empty>")
    if unknown:
        raise ValueError(f"unknown LPG curve id/code for scope {scope}: {', '.join(unknown)}")
    return resolved


def _batches(values: Sequence[str], batch_size: int) -> Iterator[tuple[str, ...]]:
    for start in range(0, len(values), batch_size):
        yield tuple(values[start:start + batch_size])


def _workbook_batch_id(batch_index: int, candidate_ids: Sequence[str]) -> str:
    if len(candidate_ids) == 1:
        return candidate_ids[0]
    digest = hashlib.sha1("|".join(candidate_ids).encode("utf-8")).hexdigest()[:8]
    return f"batch{batch_index:03d}_{digest}"


def build_backfill_workbooks(
    *,
    start_year: int,
    end_year: int,
    scope: str = "all",
    symbols: Sequence[str] | None = None,
    batch_size: int = 1,
    force: bool = False,
) -> list[Path]:
    """Build deterministic, retryable History-Symbol workbooks.

    A batch defaults to one symbol so an invalid or unentitled UDF cannot stop
    unrelated history from being saved.  ``symbols`` accepts configured Platts
    symbols or candidate ids and is intentionally scoped before batching.
    """
    _validate_scope(scope)
    if start_year > end_year:
        raise ValueError("start_year must be <= end_year")
    if batch_size < 1 or batch_size > 10:
        raise ValueError("batch_size must be between 1 and 10")
    scopes = ("asia", "overnight") if scope == "all" else (scope,)
    requested = _resolve_candidate_ids(symbols, scope=scope)
    paths: list[Path] = []
    for year in range(start_year, end_year + 1):
        for item in scopes:
            selected = [
                candidate.candidate_id for candidate in candidates_for_scope(item)
                if candidate.enabled and candidate.symbol
                and (requested is None or candidate.candidate_id in requested)
            ]
            for batch_index, batch in enumerate(_batches(selected, batch_size), 1):
                batch_id = _workbook_batch_id(batch_index, batch)
                paths.append(build_workbook(
                    workbook_path(
                        item, purpose="backfill", year=year, batch_id=batch_id,
                    ),
                    scope=item,
                    purpose="backfill",
                    year=year,
                    force=force,
                    include_discovery=False,
                    candidate_ids=batch,
                    query_kinds=("history",),
                    include_curves=False,
                    include_ewindow=False,
                ))
    return paths


def build_curve_workbooks(
    *,
    scope: str = "all",
    curve_ids: Sequence[str] | None = None,
    batch_size: int = 1,
    force: bool = False,
) -> list[Path]:
    """Build FC CurveData workbooks isolated from daily price refreshes."""
    _validate_scope(scope)
    if batch_size < 1 or batch_size > 10:
        raise ValueError("batch_size must be between 1 and 10")
    requested = _resolve_curve_ids(curve_ids, scope=scope)
    scopes = ("asia", "overnight") if scope == "all" else (scope,)
    paths: list[Path] = []
    for item in scopes:
        selected = [
            curve.candidate_id for curve in curves_for_scope(item)
            if curve.enabled and curve.curve_code
            and (requested is None or curve.candidate_id in requested)
        ]
        for batch_index, batch in enumerate(_batches(selected, batch_size), 1):
            batch_id = _workbook_batch_id(batch_index, batch)
            paths.append(build_workbook(
                workbook_path(item, purpose="curve", batch_id=batch_id),
                scope=item,
                purpose="curve",
                force=force,
                include_discovery=False,
                candidate_ids=(),
                curve_candidate_ids=batch,
                query_kinds=(),
                include_curves=True,
                include_ewindow=False,
            ))
    return paths


def build_moc_workbooks(*, scope: str = "all", force: bool = False) -> list[Path]:
    """Build isolated eWindow TradeData workbooks for the MOC view."""
    _validate_scope(scope)
    scopes = ("asia", "overnight") if scope == "all" else (scope,)
    return [
        build_workbook(
            workbook_path(item, purpose="moc"),
            scope=item,
            purpose="moc",
            force=force,
            include_discovery=False,
            candidate_ids=(),
            curve_candidate_ids=(),
            query_kinds=(),
            include_curves=False,
            include_ewindow=True,
        )
        for item in scopes
    ]


def parse_workbook(path: Path | str) -> dict[str, Any]:
    """Parse cached Add-in values into normalized records and entitlement results."""
    from openpyxl import load_workbook

    source = Path(path).resolve()
    if not source.exists():
        raise FileNotFoundError(source)
    wb = load_workbook(source, data_only=True, read_only=True)
    try:
        if "_query_manifest" not in wb.sheetnames:
            raise ValueError("workbook has no _query_manifest sheet")
        specs = _read_manifest(wb["_query_manifest"])
        runtime_status = _read_runtime_status(wb)
        discovery = _parse_discovery_sheets(wb)
        discovery_valid = any(item.get("rows") and not item.get("error_code") for item in discovery)
        session_valid = (
            str(runtime_status.get("refresh_state") or "").lower() == "success"
            and (not discovery or discovery_valid)
        )
        records: list[dict[str, Any]] = []
        entitlements: list[dict[str, Any]] = []
        query_errors: list[dict[str, str]] = []
        scopes: set[str] = set()
        seen_candidates: set[str] = set()

        for spec in specs:
            scopes.add(spec.scope)
            seen_candidates.add(spec.candidate_id)
            if spec.sheet not in wb.sheetnames:
                entitlements.append(_entitlement(spec, "error", 0, "missing_sheet"))
                query_errors.append({"query_id": spec.query_id, "reason_code": "missing_sheet"})
                continue
            rows = list(wb[spec.sheet].iter_rows(values_only=True))
            error_code = _formula_error_code(rows)
            try:
                parsed = _parse_query(rows, spec)
            except (TypeError, ValueError, OverflowError):
                parsed = []
                error_code = error_code or "parse_error"
            parsed = [_clean_record(record) for record in parsed]
            parsed = [record for record in parsed if record]
            if spec.kind == "history" and spec.year is not None:
                # This Add-in build accepts the proven lower AssessDate bound
                # but can reject an additional upper bound. Keep each annual
                # batch deterministic after retrieval instead.
                year_prefix = f"{spec.year:04d}-"
                parsed = [
                    record for record in parsed
                    if str(record.get("assess_date") or "").startswith(year_prefix)
                ]
            records.extend(parsed)
            if parsed:
                status, reason = "entitled", "data_returned"
            elif error_code in {"value_error", "reference_error"} and spec.candidate_id:
                if session_valid:
                    status, reason = "unentitled", "not_entitled_or_invalid_symbol"
                else:
                    status, reason = "error", "session_or_formula_error"
            elif error_code == "not_available":
                status, reason = "pending_review", "not_available"
            elif error_code:
                status, reason = "error", error_code
            else:
                status, reason = "pending_review", "empty_result"
            entitlements.append(_entitlement(spec, status, len(parsed), reason))
            if status == "error":
                query_errors.append({"query_id": spec.query_id, "reason_code": reason})

        concrete_scope = next(iter(scopes)) if len(scopes) == 1 else "all"
        for item in discovery:
            item["scope"] = concrete_scope
        # Only the daily/current catalog owns the broad entitlement matrix.
        # A one-symbol history retry or curve/MOC workbook must not downgrade
        # unrelated candidates merely because they were intentionally absent.
        if any(spec.kind in {"current", "correction"} for spec in specs):
            for candidate in candidates_for_scope(concrete_scope):
                if candidate.candidate_id in seen_candidates:
                    continue
                entitlements.append({
                    "canonical_key": canonical_key(candidate.candidate_id),
                    "candidate_id": candidate.candidate_id,
                    "series_id": candidate.series_id,
                    "symbol": candidate.symbol or None,
                    "curve_code": None,
                    "dataset": "MD",
                    "data_series": None,
                    "scope": candidate.scope,
                    "query_group": candidate.query_group,
                    "status": "pending_review",
                    "record_count": 0,
                    "reason_code": "symbol_not_confirmed" if not candidate.symbol else "not_queried",
                })

        manifest_years = {spec.year for spec in specs if spec.year is not None}
        kinds = {spec.kind for spec in specs}
        detected_purpose = "daily"
        if kinds == {"history"} and manifest_years:
            detected_purpose = "backfill"
        elif kinds == {"curve"}:
            detected_purpose = "curve"
        elif kinds == {"ewindow"}:
            detected_purpose = "moc"
        purpose = str(runtime_status.get("purpose") or detected_purpose).lower()
        if purpose not in VALID_PURPOSES:
            purpose = detected_purpose
        runtime_year = runtime_status.get("year")
        parsed_year = (
            int(runtime_year) if runtime_year not in (None, "") else
            next(iter(manifest_years)) if len(manifest_years) == 1 else None
        )
        return {
            "schema_version": SCHEMA_VERSION,
            "generated_at": _utc_now(),
            "scope": concrete_scope,
            "purpose": purpose,
            "year": parsed_year,
            "workbook": str(source),
            "workbook_mtime": dt.datetime.fromtimestamp(
                source.stat().st_mtime, tz=dt.timezone.utc
            ).isoformat(),
            "status": "success" if records else "discovery_only" if session_valid else "empty",
            "records": _dedupe_records(records),
            "entitlement_results": _dedupe_entitlements(entitlements),
            "runtime_status": runtime_status,
            "discovery": discovery,
            "errors": query_errors,
        }
    finally:
        wb.close()


def combine_payloads(payloads: Sequence[Mapping[str, Any]], *, requested_scope: str) -> dict[str, Any]:
    _validate_scope(requested_scope)
    records: list[dict[str, Any]] = []
    entitlements: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    workbooks: list[str] = []
    for payload in payloads:
        records.extend(payload.get("records", []))
        entitlements.extend(payload.get("entitlement_results", []))
        errors.extend(payload.get("errors", []))
        workbook = payload.get("workbook")
        if workbook:
            workbooks.append(str(workbook))
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": _utc_now(),
        "scope": requested_scope,
        "workbook": workbooks,
        "workbook_mtime": None,
        "status": "success" if records and not errors else "partial_success" if records else
                  "discovery_only" if entitlements else "empty",
        "records": _dedupe_records(records),
        "entitlement_results": _dedupe_entitlements(entitlements),
        "runtime_status": [payload.get("runtime_status") for payload in payloads
                           if payload.get("runtime_status")],
        "discovery": [item for payload in payloads for item in (payload.get("discovery") or [])],
        "errors": errors,
    }


def write_staging(
    payload: Mapping[str, Any],
    *,
    staging_dir: Path | str = STAGING_DIR,
    purpose: str | None = None,
    year: int | None = None,
    batch_id: str | None = None,
) -> Path | None:
    """Atomically stage data, inferring specialized workbook destinations."""
    runtime_status = payload.get("runtime_status")
    runtime = runtime_status if isinstance(runtime_status, Mapping) else {}
    inferred_purpose = str(
        payload.get("purpose") or runtime.get("purpose") or "daily"
    ).lower()
    purpose = str(purpose or inferred_purpose).lower()
    if purpose == "history":
        purpose = "backfill"
    if purpose not in VALID_PURPOSES:
        raise ValueError(f"purpose must be one of {sorted(VALID_PURPOSES)}")
    if year is None:
        raw_year = payload.get("year") or runtime.get("year")
        year = int(raw_year) if raw_year not in (None, "") else None
    if purpose == "backfill" and year is None:
        raise ValueError("backfill staging requires year")
    if batch_id is None and purpose != "daily":
        workbook = payload.get("workbook")
        if isinstance(workbook, (str, Path)) and str(workbook):
            batch_id = Path(workbook).stem
    destination = Path(staging_dir).resolve()
    destination.mkdir(parents=True, exist_ok=True)
    records = list(payload.get("records") or [])
    status_payload = {
        "schema_version": SCHEMA_VERSION,
        "updated_at": _utc_now(),
        "scope": payload.get("scope", "all"),
        "purpose": purpose,
        "year": year,
        "state": payload.get("status", "unknown"),
        "record_count": len(records),
        "error_count": len(payload.get("errors") or []),
        "reason_codes": sorted({
            str(item.get("reason_code")) for item in (payload.get("errors") or [])
            if item.get("reason_code")
        }),
    }
    status_dir = destination if purpose == "daily" else destination / purpose
    _atomic_json(status_dir / "status.json", status_payload)
    entitlements = list(payload.get("entitlement_results") or [])
    discovery = list(payload.get("discovery") or [])
    if not records and not entitlements and not discovery:
        return None

    clean_payload = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": payload.get("generated_at") or _utc_now(),
        "scope": payload.get("scope", "all"),
        "purpose": purpose,
        "year": year,
        "workbook": payload.get("workbook"),
        "workbook_mtime": payload.get("workbook_mtime"),
        "status": payload.get("status", "success"),
        "records": _dedupe_records(records),
        "entitlement_results": _dedupe_entitlements(entitlements),
        "runtime_status": payload.get("runtime_status") or {},
        "discovery": discovery,
        "errors": _sanitize_errors(payload.get("errors") or []),
    }

    runs_dir = destination / "runs"
    stamp = dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    batch_suffix = f"_{_safe_slug(batch_id)}" if batch_id else ""
    run_path = runs_dir / (
        f"{stamp}_{_safe_slug(purpose)}_{_safe_slug(str(clean_payload['scope']))}"
        f"{batch_suffix}.json"
    )
    _atomic_json(run_path, clean_payload)

    if purpose == "backfill":
        backfill_path = destination / "backfill" / (
            f"{year}_{_safe_slug(str(clean_payload['scope']))}{batch_suffix}.json"
        )
        _atomic_json(backfill_path, clean_payload)
        return backfill_path

    if purpose in {"curve", "moc"}:
        specialized_path = destination / purpose / (
            f"latest_{_safe_slug(str(clean_payload['scope']))}{batch_suffix}.json"
        )
        _atomic_json(specialized_path, clean_payload)
        return specialized_path

    latest_path = destination / "latest.json"
    merged = _merge_latest(latest_path, clean_payload)
    _atomic_json(latest_path, merged)
    return latest_path


def write_runtime_status(
    *,
    scope: str,
    state: str,
    reason_code: str,
    staging_dir: Path | str = STAGING_DIR,
    purpose: str = "daily",
) -> Path:
    _validate_scope(scope)
    if purpose not in VALID_PURPOSES:
        raise ValueError(f"purpose must be one of {sorted(VALID_PURPOSES)}")
    payload = {
        "schema_version": SCHEMA_VERSION,
        "updated_at": _utc_now(),
        "scope": scope,
        "state": state,
        "record_count": 0,
        "error_count": 1,
        "reason_codes": [reason_code],
    }
    destination = Path(staging_dir).resolve()
    target = (destination if purpose == "daily" else destination / purpose) / "status.json"
    _atomic_json(target, payload)
    return target


def _read_runtime_status(wb: Any) -> dict[str, Any]:
    if "_runtime_status" not in wb.sheetnames:
        return {"refresh_state": "unknown"}
    result: dict[str, Any] = {}
    for row in wb["_runtime_status"].iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            result[str(row[0])] = _json_cell(row[1] if len(row) > 1 else None)
    return result or {"refresh_state": "unknown"}


def _parse_discovery_sheets(wb: Any, max_rows: int = 500,
                            max_columns: int = 60) -> list[dict[str, Any]]:
    """Retain Add-in catalog/schema outputs inside the private staging payload."""
    output: list[dict[str, Any]] = []
    for name in DISCOVERY_SHEETS:
        if name not in wb.sheetnames:
            continue
        rows: list[list[Any]] = []
        raw_rows: list[tuple[Any, ...]] = []
        for index, row in enumerate(wb[name].iter_rows(min_row=2, values_only=True)):
            if index >= max_rows:
                break
            values = list(row[:max_columns])
            while values and values[-1] is None:
                values.pop()
            if values:
                clean = [_json_cell(value) for value in values]
                rows.append(clean)
                raw_rows.append(tuple(values))
        if rows:
            output.append({
                "sheet": name,
                "rows": rows,
                "row_count": len(rows),
                "truncated": len(rows) >= max_rows,
                "error_code": _formula_error_code(raw_rows),
            })
    return output


def _json_cell(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _read_manifest(ws: Any) -> list[QuerySpec]:
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    names = [str(item or "") for item in headers]
    specs: list[QuerySpec] = []
    for row in rows:
        values = dict(zip(names, row))
        if not values.get("query_id"):
            continue
        year = values.get("year")
        specs.append(QuerySpec(
            query_id=str(values.get("query_id") or ""),
            sheet=str(values.get("sheet") or ""),
            candidate_id=str(values.get("candidate_id") or ""),
            canonical_key=str(values.get("canonical_key") or
                              canonical_key(str(values.get("candidate_id") or ""))),
            series_id=str(values.get("series_id") or ""),
            label=str(values.get("label") or ""),
            dataset=str(values.get("dataset") or ""),
            data_series=str(values.get("data_series") or ""),
            kind=str(values.get("kind") or ""),
            scope=str(values.get("scope") or ""),
            query_group=str(values.get("query_group") or ""),
            symbol=str(values.get("symbol") or ""),
            curve_code=str(values.get("curve_code") or ""),
            basis=str(values.get("basis") or ""),
            currency=str(values.get("currency") or ""),
            uom=str(values.get("uom") or ""),
            year=int(year) if year not in (None, "") else None,
        ))
    return specs


def _parse_query(rows: Sequence[Sequence[Any]], spec: QuerySpec) -> list[dict[str, Any]]:
    if spec.kind in {"current", "history", "correction"}:
        return _parse_market_data(rows, spec)
    if spec.kind == "curve":
        return _parse_curve_data(rows, spec)
    if spec.kind == "ewindow":
        return _parse_ewindow(rows, spec)
    return []


def _parse_market_data(rows: Sequence[Sequence[Any]], spec: QuerySpec) -> list[dict[str, Any]]:
    table = _find_header(rows, required_any={"value", "price"}, date_headers={"assessdate", "date"})
    if table:
        header_idx, columns = table
        records: list[dict[str, Any]] = []
        for row in rows[header_idx + 1:]:
            value = _number(_row_value(row, columns, "value", "price"))
            assess = _iso_date(_row_value(row, columns, "assessdate", "date"))
            if value is None or not assess:
                continue
            symbol = str(_row_value(row, columns, "symbol") or spec.symbol).strip().upper()
            bate = str(_row_value(row, columns, "bate") or "c").strip().lower()
            records.append(_base_record(spec, {
                "record_type": "correction" if spec.kind == "correction" else "price_observation",
                "symbol": symbol,
                "description": _text(_row_value(row, columns, "description")) or spec.label,
                "value": value,
                "currency": _text(_row_value(row, columns, "currency")) or spec.currency,
                "uom": _text(_row_value(row, columns, "uom", "unit")) or spec.uom,
                "bate": bate,
                "assess_date": assess,
                "mod_date": _iso_datetime(_row_value(row, columns, "moddate", "modifieddate")),
                "is_corrected": _boolish(_row_value(row, columns, "iscorrected")),
                "correction_type": _text(_row_value(row, columns, "correctiontype")),
            }))
        if records:
            return records

    return _parse_market_pivot(rows, spec)


def _parse_market_pivot(rows: Sequence[Sequence[Any]], spec: QuerySpec) -> list[dict[str, Any]]:
    symbol_row = None
    symbol_columns: list[int] = []
    expected = spec.symbol.upper()
    for idx, row in enumerate(rows[:40]):
        hits = [col for col, value in enumerate(row) if str(value or "").strip().upper() == expected]
        if hits:
            symbol_row, symbol_columns = idx, hits
            break
    if symbol_row is None:
        return []

    bate_row: Sequence[Any] | None = None
    for row in rows[symbol_row + 1:symbol_row + 10]:
        first = _text(row[0] if row else None).lower()
        if "date" in first:
            bate_row = row
            break

    description = spec.label
    currency = spec.currency
    uom = spec.uom
    for row in rows[symbol_row + 1:symbol_row + 10]:
        for col in symbol_columns:
            value = _text(row[col] if col < len(row) else None)
            upper = value.upper()
            if upper in {"USD", "USC", "EUR", "GBP", "CNY", "JPY"}:
                currency = upper
            elif upper in {"MT", "BBL", "GAL", "MMB", "DAY", "LOT"}:
                uom = upper
            elif value and len(value) > 12 and "date" not in value.lower():
                description = value

    records: list[dict[str, Any]] = []
    for row in rows[symbol_row + 1:]:
        assess = next((_iso_date(row[col]) for col in range(min(2, len(row))) if _iso_date(row[col])), None)
        if not assess:
            continue
        for col in sorted(symbol_columns, key=lambda item: _bate_priority(_cell(bate_row, item))):
            value = _number(row[col] if col < len(row) else None)
            if value is None:
                continue
            bate = _text(_cell(bate_row, col)).lower() or "c"
            records.append(_base_record(spec, {
                "record_type": "correction" if spec.kind == "correction" else "price_observation",
                "symbol": expected,
                "description": description,
                "value": value,
                "currency": currency,
                "uom": uom,
                "bate": bate,
                "assess_date": assess,
                "mod_date": None,
                "is_corrected": spec.kind == "correction",
                "correction_type": "update" if spec.kind == "correction" else None,
            }))
    return records


def _parse_curve_data(rows: Sequence[Sequence[Any]], spec: QuerySpec) -> list[dict[str, Any]]:
    table = _find_header(rows, required_any={"value", "price"}, date_headers={"date", "assessdate", "valuedate"})
    records: list[dict[str, Any]] = []
    if table:
        header_idx, columns = table
        for row in rows[header_idx + 1:]:
            value = _number(_row_value(row, columns, "value", "price"))
            assess = _iso_date(_row_value(row, columns, "date", "assessdate", "valuedate"))
            if value is None or not assess:
                continue
            contract_code = _text(_row_value(row, columns, "contractcode"))
            records.append(_base_record(spec, {
                "record_type": "curve_point",
                "symbol": _text(_row_value(row, columns, "symbol")) or contract_code or None,
                "curve_code": _text(_row_value(row, columns, "curvecode")) or spec.curve_code,
                "description": _text(_row_value(row, columns, "description")) or spec.label,
                "value": value,
                "currency": _text(_row_value(row, columns, "currency")) or spec.currency,
                "uom": _text(_row_value(row, columns, "uom", "unit")) or spec.uom,
                "bate": _text(_row_value(row, columns, "bate")).lower() or "c",
                "assess_date": assess,
                "mod_date": _iso_datetime(_row_value(row, columns, "moddate", "modifieddate")),
                "contract_label": _text(_row_value(
                    row, columns, "contract", "contractcode", "tenor", "period", "strip",
                )) or None,
                "contract_date": _iso_date(_row_value(row, columns, "contractdate", "deliverydate")),
            }))
        if records:
            return records

    # Pivot fallback: first date column is valuation date, numeric columns are contracts.
    first_data = next((idx for idx, row in enumerate(rows) if row and _iso_date(row[0]) and any(_number(v) is not None for v in row[1:])), None)
    if first_data is None:
        return []
    header_row = next((rows[idx] for idx in range(first_data - 1, -1, -1) if any(_text(v) for v in rows[idx][1:])), ())
    for row in rows[first_data:]:
        assess = _iso_date(row[0] if row else None)
        if not assess:
            continue
        for col, raw in enumerate(row[1:], 1):
            value = _number(raw)
            if value is None:
                continue
            contract = _text(header_row[col] if col < len(header_row) else None)
            symbol = contract if SYMBOL_RE.match(contract.upper()) else None
            records.append(_base_record(spec, {
                "record_type": "curve_point",
                "symbol": symbol,
                "curve_code": spec.curve_code,
                "description": spec.label,
                "value": value,
                "currency": spec.currency,
                "uom": spec.uom,
                "bate": "c",
                "assess_date": assess,
                "mod_date": None,
                "contract_label": contract or f"column_{col}",
                "contract_date": None,
            }))
    return records


def _parse_ewindow(rows: Sequence[Sequence[Any]], spec: QuerySpec) -> list[dict[str, Any]]:
    table = _find_header(rows, required_any={"price", "value"}, date_headers={"orderdate", "ordertime", "updatedate", "updatetime"})
    if not table:
        return []
    header_idx, columns = table
    records: list[dict[str, Any]] = []
    for row in rows[header_idx + 1:]:
        price = _number(_row_value(row, columns, "price", "value"))
        order_date = _iso_date(_row_value(row, columns, "orderdate", "ordertime"))
        if price is None or not order_date:
            continue
        product = _text(_row_value(row, columns, "product"))
        market = _text(_row_value(row, columns, "market"))
        records.append(_base_record(spec, {
            "record_type": "ewindow_trade",
            "symbol": _text(_row_value(row, columns, "symbol")) or None,
            "description": " | ".join(item for item in (market, product) if item) or spec.label,
            "value": price,
            "currency": _text(_row_value(row, columns, "currency")) or None,
            "uom": _text(_row_value(row, columns, "uom", "unit")) or None,
            "bate": None,
            "assess_date": order_date,
            "mod_date": _iso_datetime(_row_value(row, columns, "updatetime", "updatedate")),
            "order_id": _text(_row_value(row, columns, "orderid")) or None,
            "order_time": _iso_datetime(_row_value(row, columns, "ordertime")),
            "market": market or None,
            "product": product or None,
            "hub": _text(_row_value(row, columns, "hub")) or None,
            "strip": _text(_row_value(row, columns, "strip")) or None,
            "volume": _number(_row_value(row, columns, "volume", "quantity")),
            "buyer": _text(_row_value(row, columns, "buyer")) or None,
            "seller": _text(_row_value(row, columns, "seller")) or None,
        }))
    return records


def _base_record(spec: QuerySpec, values: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "source": "platts_excel",
        "series_id": spec.series_id,
        "canonical_key": spec.canonical_key,
        "candidate_id": spec.candidate_id,
        "dataset": spec.dataset,
        "data_series": spec.data_series,
        "scope": spec.scope,
        "query_group": spec.query_group,
        "basis": spec.basis or None,
        **values,
    }


def _find_header(
    rows: Sequence[Sequence[Any]], *, required_any: set[str], date_headers: set[str]
) -> tuple[int, dict[str, int]] | None:
    best: tuple[int, dict[str, int], int] | None = None
    for idx, row in enumerate(rows[:40]):
        columns: dict[str, int] = {}
        for col, value in enumerate(row):
            key = _norm_header(value)
            if key:
                columns.setdefault(key, col)
        score = len(set(columns) & (required_any | date_headers | {
            "symbol", "description", "bate", "currency", "uom", "market", "product",
            "curvecode", "contract", "orderid", "hub", "strip",
        }))
        if set(columns) & required_any and set(columns) & date_headers and (best is None or score > best[2]):
            best = (idx, columns, score)
    return (best[0], best[1]) if best else None


def _row_value(row: Sequence[Any], columns: Mapping[str, int], *names: str) -> Any:
    for name in names:
        col = columns.get(name)
        if col is not None and col < len(row):
            return row[col]
    return None


def _formula_error_code(rows: Sequence[Sequence[Any]]) -> str | None:
    errors = {_text(value).upper() for row in rows[:80] for value in row if _text(value).upper() in ERROR_VALUES}
    if "#VALUE!" in errors or "#NAME?" in errors:
        return "value_error"
    if "#N/A" in errors:
        return "not_available"
    if "#REF!" in errors:
        return "reference_error"
    if errors:
        return "formula_error"
    return None


def _entitlement(spec: QuerySpec, status: str, count: int, reason: str) -> dict[str, Any]:
    return {
        "canonical_key": spec.canonical_key,
        "candidate_id": spec.candidate_id,
        "series_id": spec.series_id,
        "symbol": spec.symbol or None,
        "curve_code": spec.curve_code or None,
        "dataset": spec.dataset,
        "data_series": spec.data_series,
        "scope": spec.scope,
        "query_group": spec.query_group,
        "status": status,
        "record_count": count,
        "reason_code": reason,
    }


def _current_formula(symbol: str) -> str:
    return f'=PlattsGetData("MD","Current-Symbol",,"Symbol=\'{symbol}\'")'


def _history_formula(symbol: str, *, year: int | None = None) -> str:
    if year is not None:
        filter_text = (
            f"Symbol in ('{symbol}') and AssessDate>={year}-01-01 and "
            "Bate in ('c','u')"
        )
        return f'=PlattsGetData("MD","History-Symbol",,"{filter_text}")'
    return (
        '=PlattsGetData("MD","History-Symbol",,"Symbol in (\'' + symbol
        + "') and AssessDate>=" + '"&TEXT(PlattsToday()-45,"yyyy-mm-dd")&"'
        + " and Bate in ('c','u')"
        + '")'
    )


def _correction_formula(symbol: str) -> str:
    return (
        '=PlattsGetData("MD","Correction-Symbol",,"Symbol=\'' + symbol
        + "' and ModDate>=" + '"&TEXT(PlattsToday()-45,"yyyy-mm-dd")&"'
        + " and Bate in ('c','u','e','h','l')" + '")'
    )


def _curve_formula(curve_code: str, *, year: int | None = None) -> str:
    if year is not None:
        filter_text = f"CurveCode='{curve_code}' and Date>={year}-01-01 and Date<={year}-12-31"
        return f'=PlattsGetData("FC","CurveData",,"{filter_text}")'
    return (
        '=PlattsGetData("FC","CurveData",,"CurveCode=\'' + curve_code
        + "' and Date>=" + '"&TEXT(PlattsToday()-10,"yyyy-mm-dd")&"'
        + '")'
    )


def _ewindow_filters(scope: str, *, purpose: str, year: int | None) -> tuple[tuple[str, str], ...]:
    if purpose == "backfill":
        date_filter = f"Order_Date>={year}-01-01 and Order_Date<={year}-12-31"
    else:
        date_filter = 'Order_Date>="&TEXT(PlattsToday()-14,"yyyy-mm-dd")&"'
    product_filter = "(Market='*LPG*' or Product='*Propane*' or Product='*Butane*')"
    if scope == "asia":
        return (("asia_lpg", f"Window_Region='asia' and {date_filter} and {product_filter}"),)
    return ((
        "americas_ngl",
        f"Window_Region='americas' and {date_filter} and {product_filter}",
    ),)


def _ewindow_formula(filter_text: str) -> str:
    # filter_text deliberately contains Excel string concatenation around the
    # rolling date, matching the working MD History-Symbol formula pattern.
    return f'=PlattsGetData("eWMD","TradeData",,"{filter_text}")'


def _merge_latest(path: Path, incoming: Mapping[str, Any]) -> dict[str, Any]:
    previous: dict[str, Any] = {}
    if path.exists():
        try:
            previous = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            previous = {}
    refreshed_scopes = {str(record.get("scope")) for record in incoming.get("records", []) if record.get("scope")}
    retained_records = [
        record for record in previous.get("records", []) if str(record.get("scope")) not in refreshed_scopes
    ]
    incoming_keys = {
        (str(item.get("candidate_id")), str(item.get("dataset")), str(item.get("data_series")))
        for item in incoming.get("entitlement_results", [])
    }
    retained_entitlements = [
        item for item in previous.get("entitlement_results", [])
        if (str(item.get("candidate_id")), str(item.get("dataset")), str(item.get("data_series"))) not in incoming_keys
    ]
    refreshed_discovery_scopes = {
        str(item.get("scope")) for item in incoming.get("discovery", []) if item.get("scope")
    }
    retained_discovery = [
        item for item in previous.get("discovery", [])
        if str(item.get("scope")) not in refreshed_discovery_scopes
    ]
    return {
        **incoming,
        "scope": "all" if len(refreshed_scopes | {
            str(record.get("scope")) for record in retained_records if record.get("scope")
        }) > 1 else incoming.get("scope"),
        "records": _dedupe_records([*retained_records, *incoming.get("records", [])]),
        "entitlement_results": _dedupe_entitlements([
            *retained_entitlements, *incoming.get("entitlement_results", [])
        ]),
        "discovery": [*retained_discovery, *incoming.get("discovery", [])],
    }


def _dedupe_records(records: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result: dict[tuple[Any, ...], dict[str, Any]] = {}
    for raw in records:
        record = dict(raw)
        record_type = str(record.get("record_type") or "price_observation")
        if record_type in {"price_observation", "correction"}:
            # Current/History/Correction commonly contain the same assessment.
            # Collapse them before SQLite so transport metadata cannot create a
            # false revision.  Later correction queries deliberately win.
            key = ("price", record.get("series_id"), record.get("symbol"),
                   record.get("assess_date"), record.get("bate"))
        elif record_type == "curve_point":
            key = ("curve", record.get("series_id"), record.get("curve_code"),
                   record.get("assess_date"), record.get("contract_label"),
                   record.get("contract_date"))
        else:
            key = ("dataset", record.get("dataset"), record.get("order_id"),
                   record.get("order_time"), record.get("assess_date"),
                   record.get("symbol"), record.get("value"))
        result[key] = record
    return sorted(result.values(), key=lambda item: (
        str(item.get("scope") or ""), str(item.get("series_id") or ""),
        str(item.get("assess_date") or ""), str(item.get("contract_label") or ""),
        str(item.get("bate") or ""),
    ))


def _dedupe_entitlements(items: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    priority = {"entitled": 4, "unentitled": 3, "error": 2, "pending_review": 1, "retired": 0}
    result: dict[tuple[str, str, str], dict[str, Any]] = {}
    for raw in items:
        item = dict(raw)
        key = (str(item.get("candidate_id")), str(item.get("dataset")), str(item.get("data_series")))
        current = result.get(key)
        if current is None or priority.get(str(item.get("status")), -1) >= priority.get(str(current.get("status")), -1):
            result[key] = item
    return sorted(result.values(), key=lambda item: (
        str(item.get("scope") or ""), str(item.get("candidate_id") or ""),
        str(item.get("dataset") or ""), str(item.get("data_series") or ""),
    ))


def _clean_record(record: Mapping[str, Any]) -> dict[str, Any]:
    cleaned: dict[str, Any] = {}
    for key, value in record.items():
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            value = None
        if isinstance(value, (dt.date, dt.datetime)):
            value = value.isoformat()
        if isinstance(value, str):
            value = value.strip()
            if len(value) > 500:
                value = value[:500]
        cleaned[key] = value
    return cleaned


def _sanitize_errors(errors: Iterable[Mapping[str, Any]]) -> list[dict[str, str]]:
    allowed_reasons = {
        "missing_sheet", "parse_error", "value_error", "not_available", "reference_error",
        "formula_error", "empty_result", "session_expired", "excel_busy", "refresh_failed",
        "session_or_formula_error", "not_entitled_or_invalid_symbol", "not_available",
    }
    result: list[dict[str, str]] = []
    for error in errors:
        reason = str(error.get("reason_code") or "refresh_failed")
        result.append({
            "query_id": _safe_slug(str(error.get("query_id") or "unknown")),
            "reason_code": reason if reason in allowed_reasons else "refresh_failed",
        })
    return result


def _atomic_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(handle, "w", encoding="utf-8", newline="\n") as stream:
            json.dump(payload, stream, ensure_ascii=False, indent=2, sort_keys=True)
            stream.write("\n")
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    except BaseException:
        try:
            os.unlink(temporary)
        except OSError:
            pass
        raise


def _iso_date(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool) and 1 <= float(value) <= 3_000_000:
        try:
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(value))).date().isoformat()
        except (OverflowError, ValueError):
            return None
    text = _text(value)
    if not text or text.upper() in ERROR_VALUES:
        return None
    text = text.split("T", 1)[0].split(" ", 1)[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _iso_datetime(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time()).isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(value))).isoformat()
        except (OverflowError, ValueError):
            return None
    text = _text(value)
    if not text or text.upper() in ERROR_VALUES:
        return None
    normalized = text.replace("Z", "+00:00")
    try:
        return dt.datetime.fromisoformat(normalized).isoformat()
    except ValueError:
        date = _iso_date(text)
        return f"{date}T00:00:00" if date else None


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return round(number, 10) if math.isfinite(number) else None
    text = _text(value).replace(",", "")
    if not text or text.upper() in ERROR_VALUES:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return round(number, 10) if math.isfinite(number) else None


def _boolish(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    text = _text(value).lower()
    if text in {"y", "yes", "true", "1"}:
        return True
    if text in {"n", "no", "false", "0"}:
        return False
    return None


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", _text(value).lower())


def _cell(row: Sequence[Any] | None, col: int) -> Any:
    return row[col] if row is not None and col < len(row) else None


def _bate_priority(value: Any) -> int:
    return BATE_ORDER.get(_text(value).lower(), 9)


def _safe_slug(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9_.-]+", "_", value).strip("._")
    if not slug:
        slug = hashlib.sha256(value.encode("utf-8")).hexdigest()[:12]
    return slug[:100]


def _utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def _validate_scope(scope: str) -> None:
    if scope not in VALID_SCOPES:
        raise ValueError(f"scope must be one of {sorted(VALID_SCOPES)}")


__all__ = [
    "CANDIDATES", "CURVE_CANDIDATES", "PRIVATE_DIR", "SCHEMA_VERSION", "STAGING_DIR",
    "build_backfill_workbooks", "build_curve_workbooks", "build_moc_workbooks",
    "build_scope_workbooks", "build_workbook", "candidates_for_scope", "combine_payloads",
    "curves_for_scope", "parse_workbook", "workbook_path", "write_runtime_status",
    "write_staging",
]
